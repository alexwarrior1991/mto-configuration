#!/usr/bin/env python3
"""Construye data/lov-master.xlsx a partir de los workbooks de Execution Package.

Por que este script vive fuera de la aplicacion
-----------------------------------------------
Los workbooks de origen son heterogeneos y sucios: cabeceras en ES y EN, categorias
que cambian de nombre entre ficheros, notas escritas en la columna equivocada,
celdas multivalor, y hojas con formato aplicado al millon de filas (EP9A declara
2.127.071 filas fisicas). Meter todo eso en el codigo de produccion significaria
arrastrar un mapa de excepciones por fichero y un riesgo de memoria real.

En su lugar, este generador consolida los workbooks UNA vez y produce un maestro
limpio y estable de 12 columnas. La aplicacion solo sabe leer ese maestro.

Regla de oro: nunca descartar en silencio
-----------------------------------------
Todo lo que el script no sepa mapear (categoria, seccion, cabecera, hoja) va a la
hoja NO_RECONOCIDO y el proceso termina con codigo != 0. Es lo que destapo 13
categorias (~107 filas) presentes solo en EP9B/EP14A/EP14B/EP15, entre ellas los
57 codigos de POSTES SIMPLES/DOBLES/ESPECIALES que se habrian perdido sin avisar.

Uso
---
    python3 data/tools/build_lov_master.py [carpeta_workbooks] [-o salida.xlsx]

Por defecto lee data/workbook/ y escribe data/lov-master.xlsx.
"""

from __future__ import annotations

import argparse
import collections
import glob
import os
import re
import sys
import unicodedata

import openpyxl
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# La hoja BOQ buena se llama exactamente asi. EP15 trae ademas "Recuento Conjuntos-OLD"
# y ocho ficheros traen "Recuento Especiales": ninguna de las dos debe leerse.
BOQ_SHEET = "Recuento Conjuntos"
LEGEND_SHEET = "Legend"
TRACK_PREFIX = "HR TRACK"

# Tope de filas por hoja. Protege de las hojas con formato aplicado a toda la
# cuadricula (EP9A/sheet3 declara 1.048.576 filas). La hoja Track mas larga que
# hemos medido tiene ~5.200 filas reales.
MAX_ROWS_TRACK = 20_000
MAX_ROWS_BOQ = 5_000
MAX_COLS = 140

# Primera fila con datos en cada tipo de hoja (1-based).
BOQ_HEADER_ROW = 2
BOQ_FIRST_DATA_ROW = 3
TRACK_HEADER_ROW = 2
TRACK_FIRST_DATA_ROW = 4

# Valores que aparecen en las celdas de las hojas Track y que no son codigos:
# marcas de columna vacia, subcabeceras de la fila 3 y errores de formula.
TRACK_NOISE = {
    "0", "-", "TRUE", "FALSE", "P", "Ü",
    "M1", "M2", "M3", "D1", "D2", "D3", "H1", "H2", "H3",
    "E1", "E2", "E3", "B1", "B2", "B3", "W1", "W2", "W3", "A1", "A2", "A3",
}

MAX_CODE_LEN = 40
MAX_DESC_LEN = 200


# --------------------------------------------------------------------------------
# Utilidades de normalizacion
# --------------------------------------------------------------------------------

def squash(value) -> str:
    """Colapsa espacios y saltos de linea. Devuelve '' para None."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def norm_header(value) -> str:
    return squash(value).lower()


def norm_category(value) -> str:
    return squash(value).upper()


def is_noise(code: str) -> bool:
    """True si la celda no puede ser un codigo de LOV."""
    if not code:
        return True
    if code.upper() in TRACK_NOISE:
        return True
    if code.startswith("#"):                       # #REF!, #NAME?, #N/A
        return True
    if re.fullmatch(r"[-+]?\d+([.,]\d+)?", code):  # numeros sueltos
        return True
    if re.search(r"\d{2}:\d{2}:\d{2}", code):      # fechas serializadas
        return True
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", code):  # fechas tecleadas a mano
        return True
    return False


def is_pseudo_code(code: str) -> str | None:
    """Motivo por el que un texto del BOQ no es un codigo, o None si si lo es.

    El BOQ mezcla codigos reales con formulas de recuento como
    'DISC/NS + DISC/IO* + DISC/PP*' o '(puntas de feeder) FS-1 + 2xFS/PP-2'.
    """
    if len(code) > MAX_CODE_LEN:
        return f"longitud {len(code)} > {MAX_CODE_LEN}"
    if " + " in code or "+" in code and " " in code and len(code) > 20:
        return "expresion de suma"
    # Ningun codigo real lleva barra Y espacio a la vez: 'AnM/PL' y 'CP/TX-P/1100'
    # no tienen espacios, y '2HEB-300 V (CP)' o 'T-SIGN FOUND.' no llevan barra.
    # Cuando aparecen los dos es una enumeracion: 'RW1T-c o RW2T-SC/ RW2T-c (2x)'.
    if "/" in code and " " in code:
        return "enumeracion con /"
    if ";" in code:
        return "enumeracion con ;"
    if re.search(r"\s+o\s+", code, flags=re.IGNORECASE):
        return "enumeracion con 'o'"
    if code.startswith("("):
        return "nota entre parentesis"
    return None


def normalise_code(code: str) -> str:
    """Quita el asterisco final que el BOQ usa como 'y sus variantes'.

    Sin esto el catalogo sale duplicado: 'CP/TX-P*' y 'CP/TX-P' son el mismo
    codigo, y la fila con asterisco es justo la que trae el numero de plano.
    """
    return code[:-1].strip() if code.endswith("*") and len(code) > 1 else code


def looks_spanish(text: str) -> bool:
    """Heuristica ligera para separar descripcion ES de EN.

    El mismo encabezado 'Description'/'Descripcion' lleva texto en un idioma u otro
    segun el fichero, asi que hay que mirar el contenido, no la cabecera.
    """
    if not text:
        return False
    low = text.lower()
    markers = ("cion", "ción", "para ", " de ", " con ", " en ", "ñ", "á", "é", "í", "ó", "ú")
    return sum(1 for m in markers if m in low) >= 2


def strip_accents(text: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")


def parse_drawing_number(raw) -> tuple[int | None, str | None]:
    """'DW6901'->6901, 6110->6110. Devuelve (valor, aviso).

    La entidad JPA declara `Long drawingNumber`, pero el BOQ trae ademas '#N/A',
    '-', 'NEW', 'GAS6110' y 'DW7035-A', que no son convertibles.
    """
    text = squash(raw)
    if not text or text in {"-", "#N/A", "NEW"}:
        return None, None
    if isinstance(raw, float) and raw.is_integer():
        return int(raw), None
    match = re.fullmatch(r"(?:DW)?\s*(\d+)", text, flags=re.IGNORECASE)
    if match:
        return int(match.group(1)), None
    return None, f"n. de plano no numerico: {text!r}"


# --------------------------------------------------------------------------------
# Acumulador del catalogo
# --------------------------------------------------------------------------------

class Catalogue:
    """Acumula una fila por (entidad, codigo) fusionando las tres fuentes."""

    def __init__(self, cfg=None):
        self.cfg = cfg or {}
        self.rows: dict[tuple[str, str], dict] = {}
        self.discarded: list[dict] = []
        self.unknown: dict[tuple, dict] = {}

    def _key(self, entity: str, code: str):
        return entity, code.upper()

    def add(self, entity, code, *, source, ep, desc_es="", desc_en="",
            drawing=None, boq_category="", track_uses=0):
        code = normalise_code(squash(code))
        if not code:
            return

        # El enrutado va aqui para que valga igual venga la fila del BOQ, del
        # Legend o de una hoja Track.
        entity, reason = route_code(entity, code, self.cfg)
        if entity is None:
            self.discard(motivo=reason, entidad="", codigo=code, ep=ep,
                         detalle=boq_category)
            return

        key = self._key(entity, code)
        row = self.rows.get(key)
        if row is None:
            row = {
                "entity": entity,
                "code": code,
                "desc_es": "",
                "desc_en": "",
                "drawing": None,
                "sources": set(),
                "eps": set(),
                "boq_categories": set(),
                "track_uses": 0,
                "type": "",
                "notes": [],
            }
            self.rows[key] = row

        row["sources"].add(source)
        if ep:
            row["eps"].add(ep)
        if boq_category:
            row["boq_categories"].add(boq_category)
        row["track_uses"] += track_uses
        if drawing is not None and row["drawing"] is None:
            row["drawing"] = drawing

        # Conflicto de descripcion: gana la mas larga y se registra la descartada.
        for field, incoming in (("desc_es", desc_es), ("desc_en", desc_en)):
            incoming = squash(incoming)[:MAX_DESC_LEN]
            if not incoming:
                continue
            current = row[field]
            if not current:
                row[field] = incoming
            elif current != incoming:
                keep, drop = (current, incoming) if len(current) >= len(incoming) else (incoming, current)
                row[field] = keep
                self.discarded.append({
                    "motivo": "descripcion alternativa descartada",
                    "entidad": entity, "codigo": code, "ep": ep,
                    "detalle": drop, "conservado": keep,
                })

    def discard(self, *, motivo, entidad, codigo, ep, detalle=""):
        self.discarded.append({
            "motivo": motivo, "entidad": entidad, "codigo": codigo,
            "ep": ep, "detalle": detalle, "conservado": "",
        })

    def unrecognised(self, kind, value, ep, sheet, row_index=None):
        key = (kind, value, ep, sheet)
        entry = self.unknown.get(key)
        if entry is None:
            entry = {"tipo": kind, "valor": value, "ep": ep, "hoja": sheet,
                     "primera_fila": row_index, "apariciones": 0}
            self.unknown[key] = entry
        entry["apariciones"] += 1


# --------------------------------------------------------------------------------
# Lectores
# --------------------------------------------------------------------------------

def read_boq(ws, ep, cfg, cat: Catalogue):
    """Lee la hoja 'Recuento Conjuntos'.

    Particularidades: 'Cat.' y 'Obs.' vienen en celdas combinadas (hay que
    arrastrarlas), el codigo puede estar vacio con valor en 'Ref.', y algunos
    ficheros traen una segunda columna de descripcion en el otro idioma.
    """
    rows = list(ws.iter_rows(min_row=1, max_row=MAX_ROWS_BOQ,
                             max_col=MAX_COLS, values_only=True))
    if len(rows) < BOQ_FIRST_DATA_ROW:
        return

    header = rows[BOQ_HEADER_ROW - 1]
    desc_cols = [i for i, h in enumerate(header)
                 if norm_header(h) in {"description", "descripción", "descripcion"}]
    primary_desc = desc_cols[0] if desc_cols else 4
    secondary_desc = desc_cols[-1] if len(desc_cols) > 1 else None

    leaked = {norm_category(v) for v in cfg["obs_leaked_into_category"]}
    material = {norm_category(v) for v in cfg["material_categories"]}
    mapping = {norm_category(k): v for k, v in cfg["boq_categories"].items()}

    category = None
    for offset, row in enumerate(rows[BOQ_FIRST_DATA_ROW - 1:], start=BOQ_FIRST_DATA_ROW):
        raw_cat = norm_category(row[0]) if len(row) > 0 else ""
        if raw_cat and raw_cat not in leaked:
            category = raw_cat

        code_cell = row[2] if len(row) > 2 else None
        ref_cell = row[3] if len(row) > 3 else None
        code = squash(code_cell) or squash(ref_cell)
        if not code or not category:
            continue

        if category in material:
            continue
        entity = mapping.get(category)
        if entity is None:
            cat.unrecognised("categoria BOQ", category, ep, ws.title, offset)
            continue

        reason = is_pseudo_code(code)
        if reason:
            cat.discard(motivo=f"pseudocodigo ({reason})", entidad=entity,
                        codigo=code[:80], ep=ep)
            continue
        if is_noise(code):
            cat.discard(motivo="ruido", entidad=entity, codigo=code, ep=ep)
            continue

        desc = squash(row[primary_desc]) if len(row) > primary_desc else ""
        desc2 = squash(row[secondary_desc]) if secondary_desc is not None and len(row) > secondary_desc else ""

        es, en = "", ""
        for candidate in (desc, desc2):
            if not candidate:
                continue
            if looks_spanish(candidate):
                es = es or candidate
            else:
                en = en or candidate

        drawing, warn = parse_drawing_number(row[5] if len(row) > 5 else None)
        if warn:
            cat.discard(motivo=warn, entidad=entity, codigo=code, ep=ep)

        cat.add(entity, code, source="BOQ", ep=ep, desc_es=es, desc_en=en,
                drawing=drawing, boq_category=category)


def read_legend(ws, ep, cfg, cat: Catalogue):
    """Lee la hoja 'Legend', que va en dos bloques de columnas: 0/1 y 3/4.

    Una fila con descripcion y sin codigo abre seccion. El Legend es la unica
    fuente de Sectioning (el BOQ no trae ni una fila) y de los codigos mas usados
    de Anchorage y ReturnSupport.
    """
    rows = list(ws.iter_rows(min_row=1, max_row=200, max_col=6, values_only=True))
    sections = {norm_category(k): v for k, v in cfg["legend_sections"].items()}
    subheaders = {norm_category(s) for s in cfg["legend_subsection_headers"]}

    for desc_col, code_col in ((0, 1), (3, 4)):
        section = None
        for offset, row in enumerate(rows, start=1):
            desc = squash(row[desc_col]) if len(row) > desc_col else ""
            code = squash(row[code_col]) if len(row) > code_col else ""
            if not desc and not code:
                continue

            # Cabecera de seccion: texto sin codigo, en mayusculas o subseccion conocida.
            if desc and not code:
                upper = norm_category(desc)
                if upper in subheaders or (desc.upper() == desc and len(desc) > 2
                                           and not desc.startswith("(")):
                    if upper in sections:
                        section = sections[upper]
                    else:
                        cat.unrecognised("seccion Legend", upper, ep, ws.title, offset)
                        section = None
                    continue

            if section is None or not code:
                continue

            for piece in re.split(r"\s*,\s*", code):
                piece = squash(piece)
                if not piece or is_noise(piece):
                    continue

                target = section
                if target == "__SPLIT_POLE_SUPPORT":
                    target = "PoleType" if piece.upper().startswith("HEB") else "SupportType"

                if target in ("__TYPE_Foundation", "__TYPE_Portal"):
                    continue  # los catalogos de tipo salen de type_catalog

                # 'HEB-XXX', 'PH-XXX', 'CX' son comodines de la leyenda, no codigos.
                if "XXX" in piece.upper():
                    cat.discard(motivo="comodin del Legend", entidad=target,
                                codigo=piece, ep=ep)
                    continue
                if is_pseudo_code(piece):
                    continue

                cat.add(target, piece, source="LEGEND", ep=ep, desc_en=desc)


def read_track(ws, ep, cfg, cat: Catalogue):
    """Recorre una hoja 'HR Track' y cuenta el uso real de cada codigo.

    Las hojas Track NO crean catalogo por si solas: alimentan USOS_EN_TRACKS y
    aportan candidatos que salen del generador con ENABLED=NO y REVISAR=SI.
    """
    iterator = ws.iter_rows(min_row=TRACK_HEADER_ROW, max_row=MAX_ROWS_TRACK,
                            max_col=MAX_COLS, values_only=True)
    try:
        header = next(iterator)
    except StopIteration:
        return

    mapping = {norm_header(k): v for k, v in cfg["track_headers"].items()}
    ignored = {norm_header(v) for v in cfg["track_headers_ignored"]}

    columns: dict[str, list[int]] = {}
    for index, value in enumerate(header):
        name = norm_header(value)
        if not name:
            continue
        if name in mapping:
            columns.setdefault(mapping[name], []).append(index)
        elif name not in ignored:
            cat.unrecognised("cabecera Track", name, ep, ws.title, TRACK_HEADER_ROW)

    # 'Cantilevers' es una cabecera combinada sobre tres columnas (M1/M2/M3).
    if "CantileverType" in columns:
        base = columns["CantileverType"][0]
        columns["CantileverType"] = [base, base + 1, base + 2]

    if not columns:
        return

    counts: dict[str, collections.Counter] = collections.defaultdict(collections.Counter)
    for row in iterator:
        for entity, indexes in columns.items():
            for index in indexes:
                if index >= len(row) or row[index] is None:
                    continue
                # Una celda puede llevar varios valores separados por salto de linea.
                for piece in str(row[index]).split("\n"):
                    piece = squash(piece)
                    if piece and not is_noise(piece) and len(piece) <= MAX_CODE_LEN:
                        counts[entity][piece] += 1

    for entity, counter in counts.items():
        for code, uses in counter.items():
            cat.add(entity, code, source="TRACK", ep=ep, track_uses=uses)


# --------------------------------------------------------------------------------
# Derivacion de tipos
# --------------------------------------------------------------------------------

def route_code(entity, code, cfg):
    """Decide la entidad final de un codigo, o None si hay que descartarlo.

    Varias categorias del BOQ son cajones de sastre: ANCLAJES mezcla anclajes con
    material C-5xx, y HEADSPAN LOD STATION AND DEPOT mezcla portico con
    cimentaciones. Sin esto, esos codigos aterrizan en la entidad equivocada y
    ademas arrastran un tipo sin resolver.
    """
    upper = code.upper()
    if upper in {squash(v).upper() for v in cfg.get("code_rejections", [])}:
        return None, "marcador sin valor de catalogo"

    for rule in cfg.get("code_reassignment", []):
        if rule["from"] != entity:
            continue
        if re.search(rule["regex"], upper):
            target = rule.get("to")
            if target is None:
                return None, f"fuera de catalogo LOV (regla {rule['regex']})"
            return target, None
    return entity, None


def resolve_type(entity, code, cfg):
    overrides = cfg.get("type_overrides", {}).get(entity, {})
    upper = code.upper()
    if upper in {k.upper() for k in overrides}:
        for k, v in overrides.items():
            if k.upper() == upper:
                return v

    for rule in cfg.get("type_rules", {}).get(entity, []):
        if "suffix" in rule and upper.endswith(rule["suffix"].upper()):
            return rule["type"]
        if "prefix" in rule and any(upper.startswith(p.upper()) for p in rule["prefix"]):
            return rule["type"]
        if "equals" in rule and any(upper == e.upper() for e in rule["equals"]):
            return rule["type"]
        if "regex" in rule and re.search(rule["regex"], upper):
            return rule["type"]
    return None


TYPED_ENTITIES = {"Foundation": "FoundationType",
                  "Portal": "PortalType",
                  "AnchorageFoundation": "AnchorageFoundationType"}


# --------------------------------------------------------------------------------
# Escritura del maestro
# --------------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")

LOVS_HEADERS = ["ENTIDAD", "CODIGO", "DESCRIPCION_EN", "DESCRIPCION_ES", "TIPO",
                "N_PLANO", "ENABLED", "ORIGEN", "CATEGORIAS_BOQ", "EPS",
                "USOS_EN_TRACKS", "REVISAR"]


def style_sheet(ws, headers, widths=None):
    for column, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for index, width in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def write_master(path, cat: Catalogue, cfg, entity_order):
    wb = openpyxl.Workbook()

    # --- LEEME
    ws = wb.active
    ws.title = "LEEME"
    guide = [
        ("Catalogo maestro de LOVs", ""),
        ("", ""),
        ("Generado por", "data/tools/build_lov_master.py a partir de data/workbook/"),
        ("Como se usa", "El importador solo carga las filas con ENABLED=SI."),
        ("", ""),
        ("ENTIDAD", "Entidad LOV destino en mto-configuration."),
        ("CODIGO", "Codigo tal cual aparece en los workbooks. Maximo 40 caracteres."),
        ("DESCRIPCION_EN / _ES", "Descripcion en ingles y en espanol, si existe."),
        ("TIPO", "Solo Foundation, Portal y AnchorageFoundation. Ver hoja TIPOS."),
        ("N_PLANO", "Numero de plano. Vacio si el origen no traia uno numerico."),
        ("ENABLED", "SI / NO. Es lo unico que decide si la fila se carga en BBDD."),
        ("ORIGEN", "BOQ, LEGEND, BOQ+LEGEND o TRACK."),
        ("CATEGORIAS_BOQ", "Categoria de origen en la hoja Recuento Conjuntos."),
        ("EPS", "Execution Packages en los que aparece el codigo."),
        ("USOS_EN_TRACKS", "Veces que se usa en hojas HR Track. Sirve para juzgar"),
        ("", "si un candidato es un codigo real o una errata: RW1 tiene miles"),
        ("", "de usos, mientras que una errata suele tener uno."),
        ("REVISAR", "SI cuando hace falta decision humana."),
        ("", ""),
        ("Filas con ORIGEN=TRACK", "Salen con ENABLED=NO porque no estan en ningun"),
        ("", "catalogo curado. Para aceptarlas, poner ENABLED=SI."),
        ("", ""),
        ("Hoja NO_RECONOCIDO", "Si tiene filas, el catalogo esta INCOMPLETO: hay que"),
        ("", "anadir el alias que falte en data/tools/aliases.yml y regenerar."),
    ]
    style_sheet(ws, ["CONCEPTO", "EXPLICACION"], [26, 96])
    for row_index, (left, right) in enumerate(guide, start=2):
        ws.cell(row=row_index, column=1, value=left).font = Font(bold=bool(left) and not right)
        ws.cell(row=row_index, column=2, value=right)

    # --- LOVS
    ws = wb.create_sheet("LOVS")
    style_sheet(ws, LOVS_HEADERS, [22, 26, 60, 60, 12, 11, 9, 14, 34, 30, 15, 9])
    order = {name: index for index, name in enumerate(entity_order)}
    ordered = sorted(cat.rows.values(),
                     key=lambda r: (order.get(r["entity"], 99), r["code"].upper()))
    for row_index, row in enumerate(ordered, start=2):
        values = [
            row["entity"], row["code"], row["desc_en"], row["desc_es"], row["type"],
            row["drawing"],
            "SI" if row["enabled"] else "NO",
            "+".join(sorted(row["sources"])),
            ", ".join(sorted(row["boq_categories"])),
            ", ".join(sorted(row["eps"])),
            row["track_uses"],
            "SI" if row["revisar"] else "NO",
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            if row["revisar"]:
                cell.fill = REVIEW_FILL

    # --- TIPOS
    ws = wb.create_sheet("TIPOS")
    style_sheet(ws, ["ENTIDAD_TIPO", "CODIGO", "DESCRIPCION_EN", "ENABLED"], [26, 16, 60, 9])
    row_index = 2
    for type_entity, entries in cfg["type_catalog"].items():
        for code, description in entries.items():
            for column, value in enumerate([type_entity, code, description, "SI"], start=1):
                ws.cell(row=row_index, column=column, value=value)
            row_index += 1

    # --- USO_TRACKS
    ws = wb.create_sheet("USO_TRACKS")
    style_sheet(ws, ["COLUMNA_TRACK", "ENTIDAD_LOV"], [30, 26])
    for row_index, (header, entity) in enumerate(sorted(cfg["track_headers"].items()), start=2):
        ws.cell(row=row_index, column=1, value=header)
        ws.cell(row=row_index, column=2, value=entity)

    # --- DESCARTADOS
    ws = wb.create_sheet("DESCARTADOS")
    style_sheet(ws, ["MOTIVO", "ENTIDAD", "CODIGO", "EP", "DETALLE", "CONSERVADO"],
                [38, 22, 30, 10, 60, 60])
    for row_index, item in enumerate(cat.discarded, start=2):
        for column, key in enumerate(["motivo", "entidad", "codigo", "ep", "detalle", "conservado"], start=1):
            ws.cell(row=row_index, column=column, value=item.get(key, ""))

    # --- NO_RECONOCIDO
    ws = wb.create_sheet("NO_RECONOCIDO")
    style_sheet(ws, ["TIPO", "VALOR", "EP", "HOJA", "PRIMERA_FILA", "APARICIONES"],
                [20, 50, 10, 30, 14, 13])
    unknown = sorted(cat.unknown.values(), key=lambda u: (-u["apariciones"], u["valor"]))
    for row_index, item in enumerate(unknown, start=2):
        for column, key in enumerate(["tipo", "valor", "ep", "hoja", "primera_fila", "apariciones"], start=1):
            ws.cell(row=row_index, column=column, value=item[key])

    wb.save(path)
    return ordered, unknown


# --------------------------------------------------------------------------------
# Orquestacion
# --------------------------------------------------------------------------------

def discover(folder):
    """Todos los workbooks de la carpeta, sin depender de mayusculas.

    EP14A.XLSM y EP14B.XLSM traen la extension en mayusculas.
    """
    found = set()
    for pattern in ("*.xlsm", "*.xlsx", "*.XLSM", "*.XLSX"):
        found.update(glob.glob(os.path.join(folder, pattern)))
    return sorted(found)


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("folder", nargs="?", default="data/workbook",
                        help="carpeta con los workbooks (por defecto data/workbook)")
    parser.add_argument("-o", "--output", default="data/lov-master.xlsx")
    parser.add_argument("--aliases", default=os.path.join(os.path.dirname(__file__), "aliases.yml"))
    args = parser.parse_args()

    with open(args.aliases, encoding="utf-8") as handle:
        cfg = yaml.safe_load(handle)

    files = discover(args.folder)
    if not files:
        print(f"ERROR: no hay workbooks en {args.folder}", file=sys.stderr)
        return 2

    cat = Catalogue()
    cat.cfg = cfg
    print(f"Leyendo {len(files)} workbooks de {args.folder}\n")

    for path in files:
        ep = os.path.splitext(os.path.basename(path))[0]
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        boq_sheets = legend_sheets = track_sheets = 0
        try:
            for ws in wb.worksheets:
                name = ws.title
                if name == BOQ_SHEET:
                    read_boq(ws, ep, cfg, cat)
                    boq_sheets += 1
                elif name.upper().startswith(TRACK_PREFIX):
                    read_track(ws, ep, cfg, cat)
                    track_sheets += 1
                elif name == LEGEND_SHEET:
                    read_legend(ws, ep, cfg, cat)
                    legend_sheets += 1
        finally:
            wb.close()
        print(f"  {ep:8} BOQ={boq_sheets}  Legend={legend_sheets}  Track={track_sheets}")

    # Derivacion de tipos y decision de ENABLED / REVISAR.
    unresolved_types = 0
    for row in cat.rows.values():
        entity = row["entity"]
        only_track = row["sources"] == {"TRACK"}
        needs_type = entity in TYPED_ENTITIES
        type_code = resolve_type(entity, row["code"], cfg) if needs_type else None
        if needs_type:
            row["type"] = type_code or ""
            if type_code is None:
                unresolved_types += 1
        row["revisar"] = bool(only_track or (needs_type and type_code is None))
        row["enabled"] = not row["revisar"]

    ordered, unknown = write_master(args.output, cat, cfg, cfg["entities"])

    # Resumen.
    by_entity = collections.Counter(r["entity"] for r in ordered)
    enabled = collections.Counter(r["entity"] for r in ordered if r["enabled"])
    print(f"\n{'ENTIDAD':24}{'TOTAL':>7}{'ENABLED':>9}{'REVISAR':>9}")
    for entity in cfg["entities"]:
        total, active = by_entity.get(entity, 0), enabled.get(entity, 0)
        print(f"{entity:24}{total:>7}{active:>9}{total - active:>9}")
    print(f"{'TOTAL':24}{len(ordered):>7}{sum(enabled.values()):>9}"
          f"{len(ordered) - sum(enabled.values()):>9}")
    print(f"\nDescartados: {len(cat.discarded)}   Tipos sin resolver: {unresolved_types}")
    print(f"Escrito: {args.output}")

    if unknown:
        print(f"\nATENCION: {len(unknown)} valores NO RECONOCIDOS "
              f"({sum(u['apariciones'] for u in unknown)} apariciones).")
        print("El catalogo esta incompleto. Revisa la hoja NO_RECONOCIDO y amplia "
              "data/tools/aliases.yml:")
        for item in unknown[:15]:
            print(f"  [{item['tipo']}] {item['valor']!r} "
                  f"({item['ep']}/{item['hoja']}, x{item['apariciones']})")
        return 1

    print("\nSin valores no reconocidos: el catalogo esta completo.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
